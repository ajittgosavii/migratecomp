"""
PG&E Migration Tool Comparator — Streamlit App
Upload CloudMigrate.store, Matilda, and Concierto Excel files for side-by-side comparison.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import io

st.set_page_config(page_title="PG&E Migration Tool Comparator", layout="wide", page_icon="🔄")

# ---- Custom CSS ----
st.markdown("""
<style>
    .main-header {font-size: 2.2rem; font-weight: bold; color: #1F4E79; text-align: center; margin-bottom: 0.5rem;}
    .sub-header {font-size: 1.1rem; color: #666; text-align: center; margin-bottom: 2rem;}
    .metric-card {background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%); padding: 1.2rem;
                  border-radius: 10px; color: white; text-align: center; margin: 0.5rem 0;}
    .metric-value {font-size: 2rem; font-weight: bold;}
    .metric-label {font-size: 0.85rem; opacity: 0.9;}
    .winner-badge {background: #E2EFDA; color: #006100; padding: 0.3rem 0.8rem;
                   border-radius: 20px; font-weight: bold; display: inline-block;}
    .stTabs [data-baseweb="tab-list"] {gap: 8px;}
    .stTabs [data-baseweb="tab"] {padding: 10px 20px; font-weight: 600;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">PG&E Cloud Migration — Tool Comparator</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">415 Applications | 4,273 Servers | Upload Excel files from CloudMigrate.store, Matilda, and Concierto for side-by-side analysis</div>', unsafe_allow_html=True)


# ---- Helper Functions ----
def load_ai_factors(wb):
    """Extract AI acceleration factors from AI Assumptions tab."""
    ws = wb['AI Assumptions']
    factors = {}
    for r in range(6, 40):
        cat = ws.cell(row=r, column=1).value
        factor = ws.cell(row=r, column=2).value
        if cat and isinstance(factor, (int, float)) and 0 < factor < 1:
            factors[cat] = round((1 - factor) * 100, 1)
    return factors


def load_activities(wb):
    """Extract migration activities with days."""
    ws = wb['Migration Activities']
    activities = []
    current_phase = ""
    for r in range(6, ws.max_row + 1):
        col1 = ws.cell(row=r, column=1).value
        col2 = ws.cell(row=r, column=2).value
        if col1 and 'Phase' in str(col1):
            current_phase = str(col1).replace('[COMPLETED]', '').strip()
            continue
        if not col2 or col2 == 'TOTAL':
            continue
        aws_t = ws.cell(row=r, column=5).value
        aws_t = aws_t if isinstance(aws_t, (int, float)) else 0
        az_t = ws.cell(row=r, column=8).value
        az_t = az_t if isinstance(az_t, (int, float)) else 0
        azl_t = ws.cell(row=r, column=11).value
        azl_t = azl_t if isinstance(azl_t, (int, float)) else 0
        team = ws.cell(row=r, column=3).value or ""
        cm_activity = ws.cell(row=r, column=14).value or ""
        activities.append({
            'Phase': current_phase,
            'Activity': str(col2).replace('[DONE]', '').strip(),
            'Team': team,
            'AWS_Trad': aws_t,
            'Azure_Trad': az_t,
            'AzLocal_Trad': azl_t,
            'Tool_Activity': cm_activity,
        })
    return pd.DataFrame(activities)


def load_roles(wb):
    """Extract team roles from Cost Estimation."""
    ws = wb['Cost Estimation']
    roles = []
    for r in range(6, 35):
        team = ws.cell(row=r, column=1).value
        role = ws.cell(row=r, column=2).value
        hr = ws.cell(row=r, column=3).value
        fte = ws.cell(row=r, column=6).value
        if team and role and isinstance(hr, (int, float)) and 'TOTAL' not in str(team):
            roles.append({
                'Team': team, 'Role': role,
                'Hourly_Rate': hr, 'FTE': fte if isinstance(fte, (int, float)) else 0
            })
    return pd.DataFrame(roles)


def load_phases(wb):
    """Extract phase costs from Cost Estimation Section B."""
    ws = wb['Cost Estimation']
    phases = []
    for r in range(30, 50):
        phase = ws.cell(row=r, column=1).value
        if not phase or 'Section' in str(phase) or 'TOTAL' in str(phase):
            continue
        td = ws.cell(row=r, column=3).value
        ad = ws.cell(row=r, column=4).value
        ftes = ws.cell(row=r, column=5).value
        rate = ws.cell(row=r, column=6).value
        if isinstance(td, (int, float)):
            phases.append({
                'Phase': str(phase).replace('[COMPLETED]', '').strip(),
                'Trad_Days': td, 'AI_Days': ad if isinstance(ad, (int, float)) else 0,
                'FTEs': ftes if isinstance(ftes, (int, float)) else 0,
                'Rate': rate if isinstance(rate, (int, float)) else 0,
            })
    return pd.DataFrame(phases)


# ---- File Upload or Auto-Load Sample Data ----
import os

SAMPLE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
SAMPLE_FILES = {
    'CloudMigrate.store': 'CloudMigrate_Store.xlsx',
    'Matilda Cloud': 'Matilda_Cloud.xlsx',
    'Concierto Migrate': 'Concierto_Migrate.xlsx',
    'AWS Transform+Kiro': 'AWS_Transform_Kiro.xlsx',
}

st.sidebar.header("Data Source")
use_sample = st.sidebar.checkbox("Use pre-loaded sample data (all 4 tools)", value=True)

tools = {}

if use_sample:
    st.sidebar.success("Loading all 4 tool comparisons from sample data")
    for name, filename in SAMPLE_FILES.items():
        filepath = os.path.join(SAMPLE_DIR, filename)
        if os.path.exists(filepath):
            tools[name] = openpyxl.load_workbook(filepath)
            st.sidebar.markdown(f"  {name}")
else:
    st.sidebar.markdown("Upload your own Excel files:")
    cm_file = st.sidebar.file_uploader("CloudMigrate.store", type=['xlsx'], key='cm')
    mt_file = st.sidebar.file_uploader("Matilda Cloud", type=['xlsx'], key='mt')
    cc_file = st.sidebar.file_uploader("Concierto Migrate", type=['xlsx'], key='cc')
    aws_file = st.sidebar.file_uploader("AWS Transform+Kiro", type=['xlsx'], key='aws')

    if cm_file:
        tools['CloudMigrate.store'] = openpyxl.load_workbook(io.BytesIO(cm_file.read()))
        cm_file.seek(0)
    if mt_file:
        tools['Matilda Cloud'] = openpyxl.load_workbook(io.BytesIO(mt_file.read()))
        mt_file.seek(0)
    if cc_file:
        tools['Concierto Migrate'] = openpyxl.load_workbook(io.BytesIO(cc_file.read()))
        cc_file.seek(0)
    if aws_file:
        tools['AWS Transform+Kiro'] = openpyxl.load_workbook(io.BytesIO(aws_file.read()))
        aws_file.seek(0)

if not tools:
    st.info("Please upload at least one Excel file or enable sample data to begin.")
    st.stop()

# ---- Parse all data ----
all_factors = {}
all_activities = {}
all_roles = {}
all_phases = {}

for name, wb in tools.items():
    all_factors[name] = load_ai_factors(wb)
    all_activities[name] = load_activities(wb)
    all_roles[name] = load_roles(wb)
    all_phases[name] = load_phases(wb)

tool_names = list(tools.keys())
tool_colors = {'CloudMigrate.store': '#2E75B6', 'Matilda Cloud': '#ED7D31', 'Concierto Migrate': '#70AD47', 'AWS Transform+Kiro': '#FF9900'}

# ---- TAB LAYOUT ----
tabs = st.tabs([
    "Executive Summary",
    "AI Factor Comparison",
    "Timeline & Cost",
    "Cost by Cloud",
    "Cost to PG&E",
    "Agentic AI Pillars",
    "Phase Breakdown",
    "Team & FTE View",
    "Activity Detail",
    "Agentic AI Radar",
])

# ======== TAB 1: Executive Summary ========
with tabs[0]:
    st.header("Executive Summary")

    # Compute stats for all tools
    tool_stats = {}
    for name in tool_names:
        factors = all_factors[name]
        avg_savings = sum(factors.values()) / len(factors) if factors else 0
        phases_df = all_phases[name]
        trad_total = phases_df['Trad_Days'].sum()
        ai_total = phases_df['AI_Days'].sum()
        trad_cost = (phases_df['Trad_Days'] * phases_df['FTEs'] * phases_df['Rate']).sum()
        ai_cost = (phases_df['AI_Days'] * phases_df['FTEs'] * phases_df['Rate']).sum()
        tool_stats[name] = {
            'savings': avg_savings, 'ai_days': ai_total, 'trad_days': trad_total,
            'ai_cost': ai_cost, 'trad_cost': trad_cost, 'days_saved': trad_total - ai_total,
        }

    # ---- KPI cards (compact) ----
    cols = st.columns(len(tool_names))
    for i, name in enumerate(tool_names):
        s = tool_stats[name]
        color = tool_colors.get(name, '#666')
        with cols[i]:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, {color}, {color}CC);
                        padding: 1.2rem; border-radius: 12px; color: white; text-align: center;">
                <div style="font-size: 1.1rem; font-weight: bold;">{name}</div>
                <div style="font-size: 2.2rem; font-weight: bold; margin: 0.3rem 0;">{s['savings']:.0f}%</div>
                <div style="opacity: 0.85; font-size: 0.85rem;">Savings | {int(s['ai_days'])}d | ${s['ai_cost']:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")

    # ---- "Who Wins Where" using native Streamlit components ----
    st.subheader("Who Wins Where")

    areas = [
        ('Discovery & Assessment', ['Infrastructure Discovery', 'App Dependency Mapping', 'Performance Baseline']),
        ('Database Migration', ['Database Analysis', 'Schema Conversion']),
        ('Planning & Wave', ['Wave Planning', 'TCO / Cost Modeling', 'License Analysis']),
        ('IaC / Terraform', ['Terraform/IaC Generation', 'Landing Zone Design']),
        ('Containerization', ['Container Artifact Gen', 'Containerize Execution']),
        ('Code Modernization', ['Framework Upgrades', 'Middleware Migration']),
        ('Rehost Execution', ['Rehost Execution', 'Replatform Execution']),
        ('Testing (FR + NFR)', ['FR Testing (60%)', 'NFR Testing (30%)']),
        ('Security & Compliance', ['Compliance & Security', 'Vuln Discovery Scan', 'Vuln Patch Execution']),
        ('Cost Optimization', ['Right-Sizing', 'Monitoring Setup']),
        ('Multi-Cloud Coverage', []),
    ]

    coverage_scores = {
        'CloudMigrate.store': 100, 'Matilda Cloud': 100,
        'Concierto Migrate': 100, 'AWS Transform+Kiro': 26,
    }

    # Build winner data for chart
    winner_rows = []
    for area_name, categories in areas:
        if area_name == 'Multi-Cloud Coverage':
            scores = {n: coverage_scores.get(n, 0) for n in tool_names}
        else:
            scores = {}
            for name in tool_names:
                vals = [all_factors[name].get(c, 0) for c in categories if c in all_factors[name]]
                scores[name] = sum(vals) / len(vals) if vals else 0

        winner = max(scores, key=scores.get)
        sorted_scores = sorted(scores.values(), reverse=True)
        is_tie = len(sorted_scores) > 1 and (sorted_scores[0] - sorted_scores[1]) < 2

        for name in tool_names:
            winner_rows.append({
                'Area': area_name, 'Tool': name, 'Score': round(scores.get(name, 0), 1),
                'Winner': 'TIE' if is_tie else winner,
            })

    winner_df = pd.DataFrame(winner_rows)

    # Grouped bar chart — this is clean and renders perfectly
    fig_wins = px.bar(
        winner_df, x='Area', y='Score', color='Tool', barmode='group',
        color_discrete_map=tool_colors,
        title='Savings % by Area — Who Wins Where',
        text='Score',
    )
    fig_wins.update_traces(texttemplate='%{text:.0f}%', textposition='outside', textfont_size=9)
    fig_wins.update_layout(
        height=500, xaxis_tickangle=-30,
        legend=dict(orientation='h', y=1.15, x=0.5, xanchor='center'),
        yaxis_title='Savings %', xaxis_title='',
    )
    st.plotly_chart(fig_wins, use_container_width=True)

    # Winner summary table using st.dataframe with emojis
    st.subheader("Winner by Area")
    summary_rows = []
    for area_name, categories in areas:
        if area_name == 'Multi-Cloud Coverage':
            scores = {n: coverage_scores.get(n, 0) for n in tool_names}
        else:
            scores = {}
            for name in tool_names:
                vals = [all_factors[name].get(c, 0) for c in categories if c in all_factors[name]]
                scores[name] = sum(vals) / len(vals) if vals else 0

        winner = max(scores, key=scores.get)
        sorted_s = sorted(scores.values(), reverse=True)
        is_tie = len(sorted_s) > 1 and (sorted_s[0] - sorted_s[1]) < 2
        winner_label = "TIE" if is_tie else winner

        row = {'Area': area_name, 'Winner': winner_label}
        for name in tool_names:
            s = scores.get(name, 0)
            is_w = (name == winner and not is_tie)
            row[name] = f"{s:.0f}%"
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # ---- Overall Winner ----
    multi_cloud_coverage = {
        'CloudMigrate.store': 1.0, 'Matilda Cloud': 1.0,
        'Concierto Migrate': 1.0, 'AWS Transform+Kiro': 0.26,
    }

    def effective_savings(name):
        raw = tool_stats[name]['savings']
        cov = multi_cloud_coverage.get(name, 1.0)
        return raw * cov + 59 * (1 - cov)

    best_tool = max(tool_names, key=effective_savings)
    best_eff = effective_savings(best_tool)

    st.success(f"**Best for PG&E Multi-Cloud: {best_tool}** — {best_eff:.0f}% effective savings across all 415 apps")
    st.warning("**Coverage Note:** PG&E = AWS (26%) + Azure (35%) + Azure Local (27%) + Other (12%). AWS-only tools cover 26% of workload. Effective savings weighted by workload split.")

# ======== TAB 2: AI Factor Comparison ========
with tabs[1]:
    st.header("AI Acceleration Factors — Side by Side")

    # Build comparison DataFrame
    all_cats = set()
    for f in all_factors.values():
        all_cats.update(f.keys())
    all_cats = sorted(all_cats)

    factor_df = pd.DataFrame({'Category': all_cats})
    for name in tool_names:
        factor_df[name] = factor_df['Category'].map(all_factors[name]).fillna(0)

    if len(tool_names) > 1:
        factor_df['Delta'] = factor_df[tool_names[0]] - factor_df[tool_names[1]]

    # Bar chart
    fig = go.Figure()
    for name in tool_names:
        fig.add_trace(go.Bar(
            name=name, x=factor_df['Category'], y=factor_df[name],
            marker_color=tool_colors.get(name, '#666'),
            text=factor_df[name].apply(lambda x: f'{x:.0f}%'),
            textposition='outside',
        ))
    fig.update_layout(
        title="AI Savings % by Category", barmode='group',
        yaxis_title="Savings %", height=600,
        xaxis_tickangle=-45, legend=dict(orientation='h', y=1.12),
    )
    st.plotly_chart(fig, use_container_width=True)

    # Detail table
    st.subheader("Detailed Factor Table")
    st.dataframe(factor_df, use_container_width=True, height=600)

# ======== TAB 3: Timeline & Cost ========
with tabs[2]:
    st.header("Timeline & Cost Comparison")

    col1, col2 = st.columns(2)

    with col1:
        # Timeline chart
        timeline_data = []
        for name in tool_names:
            phases_df = all_phases[name]
            trad = phases_df['Trad_Days'].sum()
            ai = phases_df['AI_Days'].sum()
            timeline_data.append({'Tool': 'Traditional', 'Days': trad, 'Source': name})
            timeline_data.append({'Tool': name, 'Days': ai, 'Source': name})

        tl_df = pd.DataFrame(timeline_data)
        # Show unique tools
        summary = []
        if tool_names:
            trad_days = all_phases[tool_names[0]]['Trad_Days'].sum()
            summary.append({'Tool': 'Traditional', 'Days': trad_days, 'Months': round(trad_days / 22, 1)})
        for name in tool_names:
            ai_days = all_phases[name]['AI_Days'].sum()
            summary.append({'Tool': name, 'Days': ai_days, 'Months': round(ai_days / 22, 1)})

        sum_df = pd.DataFrame(summary)
        colors = ['#C00000'] + [tool_colors.get(n, '#666') for n in tool_names]
        fig_t = px.bar(sum_df, x='Tool', y='Days', color='Tool',
                       color_discrete_sequence=colors,
                       text='Days', title='Remaining Duration (Days)')
        fig_t.update_traces(textposition='outside')
        fig_t.update_layout(showlegend=False, height=450)
        st.plotly_chart(fig_t, use_container_width=True)

    with col2:
        # Cost chart
        cost_summary = []
        if tool_names:
            p = all_phases[tool_names[0]]
            trad_cost = (p['Trad_Days'] * p['FTEs'] * p['Rate']).sum()
            cost_summary.append({'Tool': 'Traditional', 'Cost': trad_cost})
        for name in tool_names:
            p = all_phases[name]
            ai_cost = (p['AI_Days'] * p['FTEs'] * p['Rate']).sum()
            cost_summary.append({'Tool': name, 'Cost': ai_cost})

        cost_df = pd.DataFrame(cost_summary)
        fig_c = px.bar(cost_df, x='Tool', y='Cost', color='Tool',
                       color_discrete_sequence=colors,
                       text=cost_df['Cost'].apply(lambda x: f'${x:,.0f}'),
                       title='Phase Labor Cost ($)')
        fig_c.update_traces(textposition='outside')
        fig_c.update_layout(showlegend=False, height=450, yaxis_tickformat='$,.0f')
        st.plotly_chart(fig_c, use_container_width=True)

    # Savings comparison table
    st.subheader("Savings Summary")
    savings_rows = []
    for name in tool_names:
        p = all_phases[name]
        trad_d = p['Trad_Days'].sum()
        ai_d = p['AI_Days'].sum()
        trad_c = (p['Trad_Days'] * p['FTEs'] * p['Rate']).sum()
        ai_c = (p['AI_Days'] * p['FTEs'] * p['Rate']).sum()
        savings_rows.append({
            'Tool': name,
            'Traditional Days': int(trad_d),
            'AI Days': int(ai_d),
            'Days Saved': int(trad_d - ai_d),
            'Months Saved': round((trad_d - ai_d) / 22, 1),
            'Traditional Cost': f'${trad_c:,.0f}',
            'AI Cost': f'${ai_c:,.0f}',
            'Cost Saved': f'${trad_c - ai_c:,.0f}',
        })
    st.dataframe(pd.DataFrame(savings_rows), use_container_width=True)

# ======== TAB 3: Cost by Cloud (AWS / Azure / Azure Local) ========
with tabs[3]:
    st.header("Cost by Cloud Platform — AWS vs Azure vs Azure Local")

    st.markdown("**PG&E Workload:** AWS (107 apps, 719 servers) | Azure (145 apps, 852 servers) | Azure Local/HV (130 apps, 2702 servers)")

    # Per-cloud traditional days (from Migration Activities, Discovery done)
    AWS_TRAD = 533; AZ_TRAD = 538; AZL_TRAD = 586
    RATE = 1050

    # Per-tool factors per cloud
    cloud_factors = {
        'CloudMigrate.store':  {'AWS': 0.38, 'Azure': 0.38, 'Azure Local': 0.38, 'note': 'Full multi-cloud'},
        'Matilda Cloud':       {'AWS': 0.42, 'Azure': 0.42, 'Azure Local': 0.42, 'note': 'Full multi-cloud'},
        'Concierto Migrate':   {'AWS': 0.39, 'Azure': 0.39, 'Azure Local': 0.39, 'note': 'Full multi-cloud'},
        'AWS Transform+Kiro':  {'AWS': 0.34, 'Azure': 0.50, 'Azure Local': 0.52, 'note': 'AWS-native only; Azure/AzLocal = industry std'},
    }

    # Build comparison data
    cloud_data = []
    for name in tool_names:
        factors = cloud_factors.get(name, {'AWS': 0.45, 'Azure': 0.45, 'Azure Local': 0.45})
        for cloud, trad, apps, servers in [('AWS', AWS_TRAD, 107, 719), ('Azure', AZ_TRAD, 145, 852), ('Azure Local', AZL_TRAD, 130, 2702)]:
            f = factors.get(cloud, 0.45)
            ai_d = round(trad * f)
            ftes = round(17 * apps / 382)  # proportional FTE allocation
            y1_cost = ai_d * ftes * RATE
            y2_cost = y1_cost * 0.13
            total_2yr = y1_cost + y2_cost
            trad_cost = trad * ftes * RATE * 1.25
            cloud_data.append({
                'Tool': name, 'Cloud': cloud, 'Apps': apps, 'Servers': servers,
                'Trad Days': trad, 'AI Factor': f, 'AI Days': ai_d,
                'Days Saved': trad - ai_d, 'Savings %': round((1-f)*100),
                'FTEs': ftes, 'Y1 Cost': y1_cost, 'Y2 Cost': y2_cost,
                '2-Year Total': total_2yr, 'Traditional': trad_cost,
                'Net Savings': trad_cost - total_2yr,
            })

    cloud_df = pd.DataFrame(cloud_data)

    # Chart 1: AI Days by Cloud per Tool
    fig_cloud_days = px.bar(
        cloud_df, x='Cloud', y='AI Days', color='Tool', barmode='group',
        color_discrete_map=tool_colors,
        text='AI Days',
        title='AI-Accelerated Days by Cloud Platform',
    )
    fig_cloud_days.update_traces(textposition='outside')
    fig_cloud_days.update_layout(height=450, legend=dict(orientation='h', y=1.12))
    st.plotly_chart(fig_cloud_days, use_container_width=True)

    # Chart 2: 2-Year Cost by Cloud per Tool
    fig_cloud_cost = px.bar(
        cloud_df, x='Cloud', y='2-Year Total', color='Tool', barmode='group',
        color_discrete_map=tool_colors,
        text=cloud_df['2-Year Total'].apply(lambda x: f'${x:,.0f}'),
        title='2-Year Cost by Cloud Platform ($)',
    )
    fig_cloud_cost.update_traces(textposition='outside', textfont_size=9)
    fig_cloud_cost.update_layout(height=450, yaxis_tickformat='$,.0f', legend=dict(orientation='h', y=1.12))
    st.plotly_chart(fig_cloud_cost, use_container_width=True)

    # Detailed table
    st.subheader("Detailed Per-Cloud Breakdown")
    display_cloud = cloud_df.copy()
    display_cloud['AI Factor'] = display_cloud['AI Factor'].apply(lambda x: f'{x:.0%}')
    display_cloud['Savings %'] = display_cloud['Savings %'].apply(lambda x: f'{x}%')
    for col in ['Y1 Cost', 'Y2 Cost', '2-Year Total', 'Traditional', 'Net Savings']:
        display_cloud[col] = display_cloud[col].apply(lambda x: f'${x:,.0f}')
    st.dataframe(display_cloud, use_container_width=True, hide_index=True)

    # Grand total per tool
    st.subheader("Grand Total Across All Clouds")
    grand_cloud = cloud_df.groupby('Tool').agg({
        'AI Days': 'sum', 'Days Saved': 'sum',
        '2-Year Total': 'sum', 'Traditional': 'sum', 'Net Savings': 'sum',
    }).reset_index()
    grand_cloud['Effective Savings %'] = (grand_cloud['Net Savings'] / grand_cloud['Traditional'] * 100).round(0)

    for col in ['2-Year Total', 'Traditional', 'Net Savings']:
        grand_cloud[col] = grand_cloud[col].apply(lambda x: f'${x:,.0f}')
    grand_cloud['Effective Savings %'] = grand_cloud['Effective Savings %'].apply(lambda x: f'{x:.0f}%')
    st.dataframe(grand_cloud, use_container_width=True, hide_index=True)

    # AWS caveat
    st.warning("""
    **AWS Transform+Kiro caveat:** AWS-native tools (MGN, DMS, SCT, Q Developer) are best-in-class for
    **AWS targets (107 apps = 26%)** but do NOT cover Azure or Azure Local. For the remaining **308 apps (74%)**,
    industry-standard tools are assumed (50-52% savings), making the effective multi-cloud cost higher than the
    raw AWS-only savings suggest.
    """)

# ======== TAB 4: Cost to PG&E ========
with tabs[4]:
    st.header("Total Cost to PG&E — 2-Year Projection")

    # Build 2-year cost model per tool
    cost_2yr_rows = []
    for name in tool_names:
        p = all_phases[name]
        ai_labor = (p['AI_Days'] * p['FTEs'] * p['Rate']).sum()
        trad_labor = (p['Trad_Days'] * p['FTEs'] * p['Rate']).sum()

        # Platform license costs (annual)
        license_costs = {
            'CloudMigrate.store': 48000,
            'Matilda Cloud': 45000,
            'Concierto Migrate': 60000,  # Concierto enterprise license higher
            'AWS Transform+Kiro': 0,     # Pay-per-use (included in AWS bill)
        }

        # Infra cost (PG&E provides LZ, only app-level)
        infra_monthly = 500  # Minimal since PG&E provides foundational

        # Year 1: Migration execution
        y1_labor = ai_labor
        y1_license = license_costs.get(name, 40000)
        y1_infra = infra_monthly * 12
        y1_tooling = 25000  # DMS, Terraform Cloud, monitoring
        y1_training = 20000
        y1_cybersec = 35000
        y1_vuln_tooling = 18000
        y1_subtotal = y1_labor + y1_license + y1_infra + y1_tooling + y1_training + y1_cybersec + y1_vuln_tooling
        y1_contingency = y1_subtotal * 0.12
        y1_total = y1_subtotal + y1_contingency

        # Year 2: Steady-state
        y2_labor = ai_labor * 0.13  # 13% steady-state
        y2_license = y1_license
        y2_infra = infra_monthly * 12 * 0.85
        y2_tooling = 18000
        y2_training = 8000
        y2_cybersec = 25000
        y2_vuln_tooling = 18000
        y2_subtotal = y2_labor + y2_license + y2_infra + y2_tooling + y2_training + y2_cybersec + y2_vuln_tooling
        y2_contingency = y2_subtotal * 0.12
        y2_total = y2_subtotal + y2_contingency

        grand_total = y1_total + y2_total

        # Traditional 2-year
        trad_2yr = trad_labor * 1.25

        cost_2yr_rows.append({
            'Tool': name,
            'Y1 Labor': y1_labor, 'Y1 License': y1_license,
            'Y1 Infra': y1_infra, 'Y1 Tooling': y1_tooling,
            'Y1 Training': y1_training, 'Y1 CyberSec': y1_cybersec,
            'Y1 Vuln Tooling': y1_vuln_tooling,
            'Y1 Contingency (12%)': y1_contingency,
            'Year 1 Total': y1_total,
            'Y2 Labor (13%)': y2_labor, 'Y2 License': y2_license,
            'Y2 Other': y2_tooling + y2_training + y2_cybersec + y2_vuln_tooling + y2_infra,
            'Y2 Contingency (12%)': y2_contingency,
            'Year 2 Total': y2_total,
            '2-Year Grand Total': grand_total,
            'Traditional 2-Year': trad_2yr,
            'Net Savings ($)': trad_2yr - grand_total,
            'Net Savings (%)': (trad_2yr - grand_total) / trad_2yr * 100 if trad_2yr > 0 else 0,
        })

    cost_df = pd.DataFrame(cost_2yr_rows)

    # KPI metrics row
    st.subheader("2-Year Grand Total by Tool")
    cols = st.columns(len(tool_names))
    for i, name in enumerate(tool_names):
        row = cost_df[cost_df['Tool'] == name].iloc[0]
        color = tool_colors.get(name, '#666')
        with cols[i]:
            st.metric(
                label=name,
                value=f"${row['2-Year Grand Total']:,.0f}",
                delta=f"-${row['Net Savings ($)']:,.0f} saved",
                delta_color="normal",
            )

    # Stacked bar: Y1 vs Y2
    y1y2_data = []
    for _, row in cost_df.iterrows():
        y1y2_data.append({'Tool': row['Tool'], 'Period': 'Year 1 (Migration)', 'Cost': row['Year 1 Total']})
        y1y2_data.append({'Tool': row['Tool'], 'Period': 'Year 2 (Steady-State)', 'Cost': row['Year 2 Total']})
    # Add traditional
    if cost_2yr_rows:
        trad_y1 = cost_2yr_rows[0]['Traditional 2-Year'] * 0.8
        trad_y2 = cost_2yr_rows[0]['Traditional 2-Year'] * 0.2
        y1y2_data.append({'Tool': 'Traditional (No AI)', 'Period': 'Year 1 (Migration)', 'Cost': trad_y1})
        y1y2_data.append({'Tool': 'Traditional (No AI)', 'Period': 'Year 2 (Steady-State)', 'Cost': trad_y2})

    y1y2_df = pd.DataFrame(y1y2_data)
    fig_cost = px.bar(
        y1y2_df, x='Tool', y='Cost', color='Period', barmode='stack',
        title='2-Year Total Cost to PG&E — Year 1 vs Year 2',
        text=y1y2_df['Cost'].apply(lambda x: f'${x:,.0f}'),
        color_discrete_sequence=['#2E75B6', '#70AD47'],
    )
    fig_cost.update_layout(height=500, yaxis_tickformat='$,.0f', yaxis_title='Cost (USD)')
    fig_cost.update_traces(textposition='inside', textfont_size=10)
    st.plotly_chart(fig_cost, use_container_width=True)

    # Cost breakdown table
    st.subheader("Detailed Cost Breakdown")
    display_cols = ['Tool', 'Y1 Labor', 'Y1 License', 'Y1 Infra', 'Y1 Tooling', 'Y1 Training',
                    'Y1 CyberSec', 'Y1 Vuln Tooling', 'Y1 Contingency (12%)', 'Year 1 Total',
                    'Year 2 Total', '2-Year Grand Total', 'Traditional 2-Year', 'Net Savings ($)', 'Net Savings (%)']
    display_df = cost_df[display_cols].copy()
    for col in display_df.columns:
        if col not in ('Tool', 'Net Savings (%)'):
            display_df[col] = display_df[col].apply(lambda x: f'${x:,.0f}' if isinstance(x, (int, float)) else x)
        elif col == 'Net Savings (%)':
            display_df[col] = display_df[col].apply(lambda x: f'{x:.0f}%')
    st.dataframe(display_df, use_container_width=True)

    # Savings waterfall
    st.subheader("Net Savings vs Traditional")
    savings_fig = px.bar(
        cost_df, x='Tool', y='Net Savings ($)', color='Tool',
        color_discrete_map=tool_colors,
        text=cost_df['Net Savings ($)'].apply(lambda x: f'${x:,.0f}'),
        title='Net Savings vs Traditional Migration (2-Year)',
    )
    savings_fig.update_traces(textposition='outside')
    savings_fig.update_layout(height=400, showlegend=False, yaxis_tickformat='$,.0f', yaxis_title='Savings (USD)')
    st.plotly_chart(savings_fig, use_container_width=True)

    st.info("""
    **Cost Assumptions:**
    - PG&E provides Landing Zones, Networking, and foundational cloud infrastructure
    - Year 2 labor = 13% of Year 1 (steady-state operations/optimization)
    - AWS Transform+Kiro: $0 license (pay-per-use included in AWS bill) but covers AWS-only (26% of workload)
    - Contingency at 12% for 4,273-server migration with NERC CIP requirements
    """)

# ======== TAB 4: Agentic AI Pillars ========
with tabs[5]:
    st.header("Agentic AI Pillars — Capability Architecture")

    st.markdown("""
    The **Agentic AI** approach uses autonomous AI agents that can reason, plan, and execute migration tasks
    independently — unlike traditional rule-based automation. Below is how each tool's AI architecture compares
    across the 5 pillars of Agentic AI for cloud migration.
    """)

    # Define the 5 pillars
    pillars = [
        {
            'name': 'Autonomous AI Agents',
            'icon': 'Pillar 1',
            'description': 'AI agents that independently reason, plan, and execute migration tasks without step-by-step human guidance',
            'CloudMigrate.store': {
                'score': 10, 'detail': '10 Claude-powered agents: IAM, Cost, Compliance, Observability, Containerization, Wave Planning, DR, Performance, Data Masking, Rollback Decision',
                'examples': 'IAM Agent auto-designs roles for 415 apps | Rollback Agent makes go/no-go cutover decisions | DR Agent generates failover runbooks per wave',
            },
            'Matilda Cloud': {
                'score': 0, 'detail': 'No AI agents. Rule-based workflow automation only.',
                'examples': 'Workflows execute predefined steps. No autonomous decision-making.',
            },
            'Concierto Migrate': {
                'score': 3, 'detail': 'AI in CloudIgnite (discovery) + Modernize (5000+ rules). Not truly agentic — rule-driven, not reasoning.',
                'examples': 'CloudIgnite AI classifies workloads. Modernize applies deterministic code transformation rules.',
            },
            'AWS Transform+Kiro': {
                'score': 5, 'detail': 'Amazon Q Developer has agentic code transformation. Kiro IDE has autonomous task execution. AWS-scoped only.',
                'examples': 'Q Transform: agentic .NET porting (4x faster). Kiro: autonomous multi-file code changes with human oversight.',
            },
        },
        {
            'name': 'ML Prediction Models',
            'icon': 'Pillar 2',
            'description': 'Trained ML models that predict migration outcomes, classify apps, and detect anomalies',
            'CloudMigrate.store': {
                'score': 9, 'detail': '4 trained models: App Classifier (91% accuracy, 14 features), Right-Sizer, Duration Predictor, Anomaly Detector',
                'examples': '91% accuracy classifying 415 apps into 6R | Predicts per-wave duration for project scheduling | Catches post-migration anomalies before users report',
            },
            'Matilda Cloud': {
                'score': 0, 'detail': 'No ML models. Uses static rules for classification and sizing.',
                'examples': 'Rule-based app classification. No predictive capabilities.',
            },
            'Concierto Migrate': {
                'score': 4, 'detail': 'Maximize module has ML for cost optimization. CloudIgnite has AI workload recommendations.',
                'examples': 'Maximize: real-time cost forecasting. CloudIgnite: AI-powered per-workload cloud recommendation.',
            },
            'AWS Transform+Kiro': {
                'score': 2, 'detail': 'No dedicated migration ML models. Relies on Q Developer ML (general-purpose coding AI).',
                'examples': 'Compute Optimizer uses ML for sizing (AWS-only). No migration-specific prediction.',
            },
        },
        {
            'name': 'MCP Microservices',
            'icon': 'Pillar 3',
            'description': 'Dedicated microservices providing specialized intelligence for each migration domain',
            'CloudMigrate.store': {
                'score': 10, 'detail': '8 MCP services: Database Catalog, Compliance Framework, Cloud Cost, Migration Tools, Enterprise Services, Analytics, Terraform Generator, Source Scanner',
                'examples': 'Terraform Generator auto-creates HCL from discovery | DB Catalog provides cross-cloud schema mapping | Compliance Framework checks NERC CIP automatically',
            },
            'Matilda Cloud': {
                'score': 0, 'detail': 'Monolithic architecture. No dedicated microservices for migration domains.',
                'examples': 'Single application handles all functions. No specialized domain services.',
            },
            'Concierto Migrate': {
                'score': 5, 'detail': '5 modules (Migrate/Modernize/Manage/Maximize/Intelligence) but not microservice architecture.',
                'examples': 'CloudMach (execution), CloudIgnite (discovery), Maximize (cost). Modular but not independent services.',
            },
            'AWS Transform+Kiro': {
                'score': 3, 'detail': 'Uses AWS native services (MGN, DMS, SCT) as separate services. Not integrated into a migration microservice architecture.',
                'examples': 'MGN for rehost, DMS for DB, SCT for schema. Separate console/API per service — not orchestrated.',
            },
        },
        {
            'name': 'Guardrails & Safety',
            'icon': 'Pillar 4',
            'description': 'Safety controls that prevent runaway costs, unsafe operations, and data loss during AI-driven migration',
            'CloudMigrate.store': {
                'score': 10, 'detail': '5 enforced controls: Source Protection, Encryption Enforcement, Budget Guardrails ($500/action), Code Scanning, Data Validation Required',
                'examples': 'Budget cap prevents AI from spinning up expensive instances | Source protection ensures no accidental data deletion | Encryption enforced on all migrated data',
            },
            'Matilda Cloud': {
                'score': 1, 'detail': 'Basic workflow approval gates. No AI-specific guardrails.',
                'examples': 'Manual approval steps in workflows. No automated cost or safety controls.',
            },
            'Concierto Migrate': {
                'score': 3, 'detail': 'Code scanning (5000+ rules in Modernize). Zero-code interface provides transparency but not guardrails.',
                'examples': 'Modernize scans for security vulnerabilities in code. CloudMach has rollback capability. No budget guardrails.',
            },
            'AWS Transform+Kiro': {
                'score': 6, 'detail': 'AWS IAM + Config Rules + Service Control Policies. Native AWS guardrails are strong but AWS-scoped only.',
                'examples': 'IAM prevents unauthorized actions. Config Rules enforce compliance. SCPs limit regions/services. But no migration-specific guardrails.',
            },
        },
        {
            'name': 'Context & Memory',
            'icon': 'Pillar 5',
            'description': 'AI retains project context across sessions, learns from past migrations, and provides conversational assistance',
            'CloudMigrate.store': {
                'score': 10, 'detail': 'Per-project agentic memory, AI chat with SSE streaming, 39 agentic API routes, context retention across migration phases',
                'examples': 'AI remembers Phase 1 findings when executing Phase 3 | Chat assistant answers migration questions in context | Streaming responses for real-time feedback',
            },
            'Matilda Cloud': {
                'score': 0, 'detail': 'No AI memory or conversational interface. Stateless workflow execution.',
                'examples': 'Each workflow run is independent. No learning or context retention.',
            },
            'Concierto Migrate': {
                'score': 2, 'detail': 'Zero-code interface provides UI transparency. Intelligence module has data analytics. No conversational AI.',
                'examples': 'Intelligence module analyzes data patterns. No memory or chat interface.',
            },
            'AWS Transform+Kiro': {
                'score': 7, 'detail': 'Kiro IDE has project context, spec-driven development, and Claude AI chat. Q Developer retains session context.',
                'examples': 'Kiro: specs persist across sessions. Q Developer: conversational code transformation. But no migration-specific memory.',
            },
        },
    ]

    # Pillar score comparison chart
    pillar_chart_data = []
    for pillar in pillars:
        for name in tool_names:
            if name in pillar:
                pillar_chart_data.append({
                    'Pillar': pillar['name'],
                    'Tool': name,
                    'Score': pillar[name]['score'],
                })

    pillar_df = pd.DataFrame(pillar_chart_data)
    fig_pillars = px.bar(
        pillar_df, x='Pillar', y='Score', color='Tool', barmode='group',
        color_discrete_map=tool_colors,
        title='Agentic AI Pillar Scores (0-10 scale)',
        text='Score',
    )
    fig_pillars.update_traces(textposition='outside')
    fig_pillars.update_layout(
        height=500, xaxis_tickangle=-20,
        legend=dict(orientation='h', y=1.12, x=0.5, xanchor='center'),
        yaxis=dict(range=[0, 12], title='Score (0-10)'),
    )
    st.plotly_chart(fig_pillars, use_container_width=True)

    # Detailed pillar cards
    for pillar in pillars:
        st.subheader(f"{pillar['icon']}: {pillar['name']}")
        st.caption(pillar['description'])

        pcols = st.columns(len(tool_names))
        for i, name in enumerate(tool_names):
            if name in pillar:
                info = pillar[name]
                score = info['score']
                color = tool_colors.get(name, '#666')
                with pcols[i]:
                    # Score badge
                    if score >= 8:
                        st.success(f"**{name}**: {score}/10")
                    elif score >= 4:
                        st.warning(f"**{name}**: {score}/10")
                    else:
                        st.error(f"**{name}**: {score}/10")
                    st.markdown(f"**Detail:** {info['detail']}")
                    st.markdown(f"*Examples: {info['examples']}*")

    # Total pillar scores
    st.subheader("Agentic AI Total Score")
    total_scores = {}
    for name in tool_names:
        total = sum(p[name]['score'] for p in pillars if name in p)
        total_scores[name] = total

    score_df = pd.DataFrame([
        {'Tool': name, 'Total Score': score, 'Out of': 50,
         'Grade': 'A+' if score >= 40 else 'A' if score >= 30 else 'B' if score >= 20 else 'C' if score >= 10 else 'D'}
        for name, score in total_scores.items()
    ])
    st.dataframe(score_df, use_container_width=True, hide_index=True)

    best_ai = max(total_scores, key=total_scores.get)
    st.success(f"**Agentic AI Leader: {best_ai}** — {total_scores[best_ai]}/50 ({total_scores[best_ai]/50*100:.0f}%)")

    st.info("""
    **Scoring Methodology:**
    - **10/10**: Production-ready, fully autonomous, purpose-built for migration
    - **7-9**: Strong capability but with limitations (scope, coverage, maturity)
    - **4-6**: Partial capability — exists but not fully agentic or migration-specific
    - **1-3**: Minimal capability — basic rules or manual with some automation
    - **0**: No capability in this pillar
    """)

# ======== TAB 5: Phase Breakdown ========
with tabs[6]:
    st.header("Phase-by-Phase Comparison")

    for name in tool_names:
        st.subheader(name)
        p = all_phases[name].copy()
        p['Trad_Cost'] = p['Trad_Days'] * p['FTEs'] * p['Rate']
        p['AI_Cost'] = p['AI_Days'] * p['FTEs'] * p['Rate']
        p['Savings'] = p['Trad_Cost'] - p['AI_Cost']
        p['Savings_%'] = (p['Savings'] / p['Trad_Cost'] * 100).fillna(0).round(1)
        st.dataframe(p[['Phase', 'Trad_Days', 'AI_Days', 'FTEs', 'Rate', 'Trad_Cost', 'AI_Cost', 'Savings', 'Savings_%']],
                     use_container_width=True)

    # Stacked comparison
    if len(tool_names) > 1:
        st.subheader("AI Days by Phase — All Tools")
        phase_cmp = []
        for name in tool_names:
            for _, row in all_phases[name].iterrows():
                phase_cmp.append({'Tool': name, 'Phase': row['Phase'], 'AI_Days': row['AI_Days']})
        phase_cmp_df = pd.DataFrame(phase_cmp)
        fig_p = px.bar(phase_cmp_df, x='Phase', y='AI_Days', color='Tool', barmode='group',
                       color_discrete_map=tool_colors,
                       title='AI-Accelerated Days by Phase')
        fig_p.update_layout(xaxis_tickangle=-30, height=500)
        st.plotly_chart(fig_p, use_container_width=True)

# ======== TAB 5: Team & FTE ========
with tabs[7]:
    st.header("Team & FTE Distribution")

    for name in tool_names:
        st.subheader(name)
        roles_df = all_roles[name]
        team_summary = roles_df.groupby('Team').agg(
            FTEs=('FTE', 'sum'),
            Avg_Rate=('Hourly_Rate', 'mean'),
            Monthly_Cost=('FTE', lambda x: sum(x * roles_df.loc[x.index, 'Hourly_Rate'] * 8 * 22)),
        ).reset_index()
        team_summary['Annual_Cost'] = team_summary['Monthly_Cost'] * 12
        st.dataframe(team_summary, use_container_width=True)

    # FTE pie
    if tool_names:
        name = tool_names[0]
        roles_df = all_roles[name]
        team_fte = roles_df.groupby('Team')['FTE'].sum().reset_index()
        fig_fte = px.pie(team_fte, names='Team', values='FTE',
                         title=f'FTE Distribution ({name})',
                         color_discrete_sequence=px.colors.qualitative.Set2)
        fig_fte.update_traces(textinfo='label+value+percent')
        st.plotly_chart(fig_fte, use_container_width=True)

# ======== TAB 6: Activity Detail ========
with tabs[8]:
    st.header("Activity-Level Detail")

    selected_tool = st.selectbox("Select Tool", tool_names)
    acts_df = all_activities[selected_tool]

    # Filter
    phases = acts_df['Phase'].unique().tolist()
    selected_phase = st.multiselect("Filter by Phase", phases, default=phases)
    filtered = acts_df[acts_df['Phase'].isin(selected_phase)]

    st.dataframe(filtered, use_container_width=True, height=600)

    # Show tool-specific activities
    st.subheader("Tool Accelerator Activities")
    tool_acts = filtered[filtered['Tool_Activity'].str.len() > 5][['Activity', 'Team', 'Tool_Activity']]
    st.dataframe(tool_acts, use_container_width=True)

# ======== TAB 7: Agentic AI ========
with tabs[9]:
    st.header("Agentic AI Radar")

    agentic_data = {
        'Capability': [
            'Autonomous AI Agents', 'ML Prediction Models', 'MCP Microservices',
            'Guardrails & Safety', 'LLM Integration', 'Agentic API Routes',
            'AI Chat/Streaming', 'Project Memory', 'Code Scanning Rules',
            'Zero-Downtime Cutover', 'VM Throughput/Month', 'Multi-Cloud Support',
        ],
        'CloudMigrate.store': [
            '10 (Claude-powered)', '4 (91% accuracy)', '8 dedicated', '5 controls',
            'Claude AI (Anthropic)', '39 routes', 'YES (SSE)', 'YES', 'Guardrails engine',
            'Orchestrated', 'Orchestrated', 'AWS + Azure + Azure Local',
        ],
        'Matilda Cloud': [
            '0', '0', '0', '0', 'None', '0', 'NO', 'NO', 'None',
            'Workflow-based', 'Standard', 'AWS + Azure + GCP',
        ],
        'Concierto Migrate': [
            '2 (partial)', '1 (Maximize)', '3 (partial)', '1 (code scan)',
            'Not disclosed', '~10', 'Zero-code UI', 'NO', '5000+ rules (Modernize)',
            'Zero-downtime (CloudMach)', '1000+ VMs/month', 'AWS + Azure + GCP',
        ],
        'AWS Transform+Kiro': [
            'Q Developer (agentic)', '0 (uses Q ML)', '0 (native services)', 'IAM/Config native',
            'Claude AI (via Kiro)', '~15 (Q+Kiro)', 'Kiro IDE chat', 'NO', 'Q Transform 5000+',
            'MGN zero-downtime', 'MGN high-throughput', 'AWS ONLY',
        ],
    }
    ai_df = pd.DataFrame(agentic_data)

    # Show only uploaded tools
    display_cols = ['Capability'] + [n for n in tool_names if n in ai_df.columns]
    st.dataframe(ai_df[display_cols], use_container_width=True, height=500)

    # Radar chart
    if len(tool_names) >= 2:
        categories = ['AI Agents', 'ML Models', 'MCP Services', 'Guardrails',
                       'LLM', 'API Routes', 'Chat/Stream', 'Memory']
        cm_scores = [10, 4, 8, 5, 5, 5, 5, 5]
        mt_scores = [0, 0, 0, 0, 0, 0, 0, 0]
        cc_scores = [2, 1, 3, 1, 1, 2, 3, 0]

        aws_scores = [4, 1, 0, 4, 4, 3, 4, 0]  # Q Developer agentic but AWS-only

        score_map = {
            'CloudMigrate.store': cm_scores,
            'Matilda Cloud': mt_scores,
            'Concierto Migrate': cc_scores,
            'AWS Transform+Kiro': aws_scores,
        }

        fig_radar = go.Figure()
        for name in tool_names:
            if name in score_map:
                fig_radar.add_trace(go.Scatterpolar(
                    r=score_map[name] + [score_map[name][0]],
                    theta=categories + [categories[0]],
                    fill='toself', name=name,
                    line_color=tool_colors.get(name, '#666'),
                ))
        fig_radar.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 10])),
            title="Agentic AI Capability Radar",
            height=500,
        )
        st.plotly_chart(fig_radar, use_container_width=True)

    st.subheader("Key Insight")
    st.success("""
    **CloudMigrate.store** is the only platform with true Agentic AI:
    10 autonomous Claude-powered agents, 4 ML models, 8 MCP microservices, and a Guardrails engine.
    Matilda has zero AI agents (rule-based only). Concierto has partial AI in discovery + modernization
    but no autonomous agents.
    """)

# ---- Footer ----
st.divider()
st.markdown("""
<div style="text-align: center; color: #999; font-size: 0.85rem;">
    PG&E Migration Tool Comparator | 415 Apps | 4,273 Servers | Powered by Infosys Cobalt
</div>
""", unsafe_allow_html=True)
